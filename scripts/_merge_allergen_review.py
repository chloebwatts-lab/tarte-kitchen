#!/usr/bin/env python3
"""Merge rules-pass proposals + web-research results into one review set.
Writes /tmp/allergen-final-proposals.json and an xlsx review workbook.
NOTHING is written to the database here."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

proposals = json.load(open("/tmp/allergen-proposals.json"))  # rules pass, all 536

# --- web research results, keyed by EXACT original ingredient name ---
# c=contains (declared), m=may contain (trace), conf, note
R = {
 "Chocolate Veliche Belgian Dark Choc Emotion": (["SOY"], ["MILK","TREE_NUT","PEANUT","GLUTEN","WHEAT","EGG","SULPHITE"], "high", "AU label: cocoa mass/butter, soy lecithin; 'may contain milk' + facility traces"),
 "Cocoa powder, Veliche": ([], ["MILK","TREE_NUT","PEANUT","GLUTEN","WHEAT","EGG","SOY","SULPHITE"], "medium", "Sole ingredient cocoa; facility trace statement; confirm pack"),
 "Veilche Cocoa Powder": ([], ["MILK","TREE_NUT","PEANUT","GLUTEN","WHEAT","EGG","SOY","SULPHITE"], "medium", "Duplicate of Veliche cocoa powder; confirm pack"),
 "Veliche Chocolate Batons": (["SOY"], ["MILK","TREE_NUT","PEANUT","GLUTEN","WHEAT","EGG","SULPHITE"], "medium", "Soy lecithin near-certain; confirm exact label"),
 "Veliche Belgian White Choc": (["MILK","SOY"], ["TREE_NUT","PEANUT","GLUTEN","WHEAT","EGG","SULPHITE"], "high", "White couverture: milk powder + soy lecithin"),
 "Chocolate Patissier Milk Choc 34.6%": (["MILK","SOY"], ["TREE_NUT","PEANUT","WHEAT"], "high", "Milk couverture: milk powder + soya lecithin"),
 "Chocolate Dark Calypso Compound Nestle": (["MILK","SOY"], ["TREE_NUT"], "medium", "Nestle dark compound: milk solids + soy lecithin; no gluten"),
 "Chocolate Buttons White Nestle": (["MILK","SOY"], [], "medium", "Nestle white melts: whey (milk) + soy lecithin"),
 "Chocolate Lindt Piccoli White 2.5kg": (["MILK","SOY"], ["TREE_NUT"], "high", "Lindt white couverture: milk powder + soy lecithin; may contain hazelnut/almond"),
 "IRCA Chococream Pistachio 15%": (["MILK","TREE_NUT","SOY"], [], "high", "Pistachio + milk solids + soy lecithin; manufacturer declares"),
 "Nutella Chocolate Hazelnut Spread": (["MILK","TREE_NUT","SOY"], [], "high", "Hazelnut 13% + skim milk powder + soy lecithin; gluten free"),
 "Hazelnut praline": (["TREE_NUT"], ["MILK","SOY","SESAME","PEANUT"], "low", "Generic praline; only tree nut certain; verify supplier spec"),
 "Chocolate Powder": ([], ["MILK","SOY","TREE_NUT","GLUTEN","SESAME","PEANUT"], "medium", "Generic drinking choc; if milk powder added it CONTAINS milk - check label"),
 "Chocolate Sorbet": (["SOY"], ["MILK","TREE_NUT","EGG","PEANUT","GLUTEN"], "low", "Dairy-free; soy lecithin typical (flips if sunflower); brand-dependent"),
 "Fondant Soft White": ([], [], "medium", "Sugar+glucose fondant, no added allergens (Bakels); verify brand"),
 "Flan Gel Neutral Fruit Glaze": (["SULPHITE"], [], "high", "Glucose w/ preservative 220 sulphites; declares sulphites"),
 "Fresh As Raspberry - Freeze Dried": ([], [], "high", "100% raspberry, single ingredient, no allergens"),
 # asian
 "Gochujang Paste": (["SOY"], ["WHEAT","GLUTEN"], "medium", "Fermented soybean (soy); many brands list wheat - if so move to contains"),
 "Sriracha sauce": ([], [], "low", "Huy Fong = none; Flying Goose = soy+wheat. MUST check brand"),
 "Tabasco": ([], [], "high", "Pepper, vinegar, salt only - no listed allergens"),
 "Ketjap Manis (sweet Soy)": (["SOY","WHEAT","GLUTEN"], [], "high", "Soy sauce base = soy + wheat; ABC brand adds sulphite"),
 "Mirin": ([], [], "low", "Hon-mirin none; aji-mirin may share gluten lines; verify"),
 "Mirin Cooking Wine Kikkoman": ([], [], "medium", "Kikkoman aji-mirin: no listed allergens as ingredient; shared line"),
 "Miso Mayo": (["SOY","EGG"], [], "medium", "Miso=soy; egg-based mayo=egg (vegan version would drop egg)"),
 "Tamarind Paste": (["SULPHITE"], [], "medium", "Most concentrates have metabisulphite; pure pulp has none - check"),
 "Sambal Oelek": ([], ["CRUSTACEAN","FISH"], "medium", "Chilli/salt/vinegar; some add shrimp/fish - check brand"),
 "Shrimp Paste": (["CRUSTACEAN"], ["FISH"], "high", "Ground shrimp = crustacean"),
 "Togarashi": (["SESAME"], [], "high", "Shichimi blend includes sesame"),
 "Togorashi": (["SESAME"], [], "high", "Shichimi blend includes sesame (dup spelling)"),
 "Old Bay Seasoning": ([], [], "high", "Mustard+celery only (both out of FSANZ-14 scope); no others"),
 "Toasted Rice Powder": ([], [], "high", "Toasted glutinous rice - 'glutinous' is a misnomer, gluten-free"),
 # western
 "BBQ sauce": ([], [], "low", "Generic; some add soy/Worcestershire(fish) - verify brand"),
 "HP sauce": (["WHEAT","GLUTEN","SULPHITE"], [], "high", "Malt vinegar(barley)+rye flour=gluten; sulphite present"),
 "Franks Hot Sauce": ([], [], "high", "Cayenne, vinegar, water, salt, garlic - no allergens; GF"),
 "Hangover Sauce": ([], [], "low", "Could NOT identify brand - get the bottle, do not rely on this row"),
 "Sauce Red Hot Original Buffalo Wings": ([], [], "medium", "Frank's Buffalo: official=no allergens (butter flavour non-dairy); one aggregator flagged milk - confirm AU label"),
 "Worcestershire sauce": (["FISH"], ["GLUTEN"], "medium", "Anchovy=fish; gluten if malt vinegar (likely) - confirm bottle"),
 "Liquid Smoke (Wrights)": ([], [], "high", "Water + natural smoke flavour; spec lists no allergens"),
 "Demi Glaze Basic Brown GF (Maggi)": (["SOY"], ["FISH","MILK","SULPHITE"], "high", "Soy (flavours/beef fat); coeliac-endorsed GF; may contain fish/milk/sulphite"),
 "Vegetable Stock - Real Campbells": ([], [], "low", "No in-scope allergens in ingredient list (celery out of scope); confirm on-pack"),
 "Hellmans Mayonnaise": (["EGG"], [], "high", "Egg yolk = egg; GF (mustard out of scope)"),
 "Whole egg mayonnaise": (["EGG"], [], "high", "Whole egg = egg; some brands add soy - check"),
 "Vincotto": (["SULPHITE"], [], "medium", "Cooked grape must; sulphites characteristic - confirm bottle"),
 "Guacamole": ([], [], "medium", "Avocado/onion/tomato/lime - no allergens typical; verify brand"),
 "Hommus Dip (fresh)": (["SESAME"], ["TREE_NUT","MILK"], "high", "Tahini = sesame; AU products often may-contain tree nut/milk"),
 "Chimichurri - Bought In": ([], [], "low", "Herbs/oil/vinegar; wine vinegar may add sulphite - verify jar"),
 "tapenade": (["FISH"], [], "medium", "Traditional contains anchovy=fish; anchovy-free versions none; pine-nut version adds tree nut"),
 "harrissa": ([], ["SULPHITE"], "low", "Chilli/pepper/garlic base none; many add wine vinegar/lemon=sulphite; verify"),
 # baking
 "Curry Powder": ([], ["GLUTEN","WHEAT"], "medium", "Pure spice GF; shared-facility trace common - verify tin"),
 "Garam Masala": ([], ["TREE_NUT","SESAME","WHEAT","GLUTEN"], "medium", "Pure spice GF; shared spice lines - precautionary common"),
 "Italian Herb Mix": ([], ["WHEAT","GLUTEN","SESAME"], "medium", "Pure herbs GF; MasterFoods declares may-contain wheat/gluten/sesame"),
 "Bagel Seasoning": (["SESAME"], [], "high", "Everything-bagel = sesame-dominant (poppy not an FSANZ allergen)"),
 "Bread improver": (["WHEAT","GLUTEN","SOY"], [], "medium", "Wheat carrier + soy; some (Bakels Advance) also egg/milk/sulphite - check brand"),
 "Yeast - dry instant": ([], [], "high", "Yeast + emulsifier; allergen-free, GF"),
 "Nutritional yeast": ([], ["GLUTEN","SOY","SESAME","MILK","TREE_NUT","SULPHITE","PEANUT","EGG"], "medium", "Naturally none; bulk packs carry broad trace line"),
 "Isomalt": ([], [], "high", "Sugar-beet sucrose derived; no allergens, GF"),
 "Xantana": ([], ["WHEAT","SOY"], "medium", "GF in final product; fermentation substrate may be wheat/soy - verify for severe allergy"),
 "Muesli Tarte": (["GLUTEN","TREE_NUT"], ["WHEAT","SESAME","MILK","SOY","PEANUT"], "medium", "Oats=gluten (AU); nuts=tree nut. CONFIRM exact nut/seed list in house recipe"),
 "Pancake Mix Dry - Bidfood": (["WHEAT","GLUTEN"], ["MILK","EGG","SOY"], "low", "Spec behind trade login; wheat/gluten certain; milk/egg may be CONTAINS - verify pack"),
 "Pancake Mix Buttermilk": (["WHEAT","GLUTEN","MILK"], ["EGG","SOY"], "medium", "Buttermilk=milk; egg may move to contains - check label"),
 "Puffed Grain": (["GLUTEN","WHEAT"], ["PEANUT","TREE_NUT"], "low", "AMBIGUOUS: puffed wheat=wheat/gluten; rice bubbles=gluten(barley malt); pure puffed rice=NONE. Identify product"),
 # alcohol
 "Aperol": ([], [], "high", "No gluten/nuts/sulphites, vegan (bottled; RTS premix has sulphites)"),
 "Pernod": ([], [], "medium", "Distilled anise spirit; distillation removes gluten; no allergens"),
 "Triple Sec - Vok": ([], [], "medium", "Distilled orange liqueur, vegan; no declared allergens"),
 "Macadamia Liqueur (Mac by Brookie's)": (["TREE_NUT"], [], "high", "Macadamia steeped in spirit = tree nut"),
 "Mr Black Coffee Liqueur": ([], [], "high", "Distilled wheat vodka base = GF per brand; coffee/sugar/spirit"),
 "St Germain Elderflower Liqueur": ([], [], "medium", "No notifiable allergens; brand states GF"),
 "Milo": (["MILK","GLUTEN","SOY"], ["TREE_NUT","PEANUT","SESAME"], "high", "Nestle AU declares malt barley(gluten), milk solids, soy"),
 "Malt milk powder": (["MILK","GLUTEN"], [], "medium", "Malted barley(gluten)+milk; some add wheat - verify"),
 "Hot Honey": ([], [], "low", "Honey+chilli none; some brands add sesame - verify brand"),
 "Agave Syrup Senor Maguey Organic": ([], [], "high", "100% blue agave; GF, vegan, no additives"),
 "Herradura Agave Nectar": ([], [], "medium", "Single ingredient agave; no real allergens (US boilerplate ignored)"),
 # bakery/meat
 "Croissant - Mini (Bridor)": (["WHEAT","GLUTEN","MILK","EGG"], ["SOY","TREE_NUT","SESAME"], "high", "Wheat, butter/milk powder, gluten, eggs"),
 "Roti Canai Bread": (["WHEAT","GLUTEN","SOY"], ["MILK","EGG","SESAME","PEANUT","CRUSTACEAN","FISH","MOLLUSC"], "medium", "Wheat flour + palm margarine soy lecithin; broad trace line"),
 "Milk Bun (BreadTop)": (["WHEAT","GLUTEN","MILK","EGG","SOY"], [], "high", "BreadTop lists wheat, dairy, egg, soy"),
 "Butter Puff Roll - Careme": (["WHEAT","GLUTEN","MILK"], [], "high", "Wheat flour + butter(milk); no soy/egg"),
 "Butter Sheet, Tourage french": (["MILK"], [], "high", "Lamination butter = cream; dairy only"),
 "Almond Milk - UHT Barista": (["TREE_NUT"], ["SOY","GLUTEN"], "high", "Almonds = tree nut (NOT milk)"),
 "Oat Milk - UHT Barista": (["GLUTEN"], ["TREE_NUT","SOY"], "high", "Oats=gluten in AU unless certified GF (e.g. Minor Figures) - confirm brand"),
 "Ice Cream Vanilla Supreme": (["MILK"], ["EGG","SOY"], "medium", "Dairy=milk; many vanillas add egg/soy - check label"),
 "Creme Fraiche Provedores": (["MILK"], [], "high", "Cream + cultures = milk"),
 "Mortadella": (["TREE_NUT"], ["MILK","SOY"], "medium", "Classic mortadella has PISTACHIO=tree nut. If plain SKU, NONE - verify"),
 "Guanciale": ([], ["MILK"], "medium", "Pork+salt+spice+nitrite cure(not sulphite); some shared-facility milk"),
 "Jamon de Serano": ([], ["SULPHITE"], "medium", "Pork+salt, nitrate cure; most serrano has no allergen - verify"),
 "Suckling pig whole": ([], [], "high", "Plain pork - not an FSANZ allergen"),
 "Acai Mix Scoopable - Amazonia": ([], [], "medium", "Acai/cane/guarana; stated gluten/dairy/soy free"),
 "Acai Blend - Wastage allowance": ([], [], "medium", "Acai+guarana blend; free from gluten/dairy/soy"),
}

ORDER = ["MILK","EGG","FISH","CRUSTACEAN","SHELLFISH","MOLLUSC","PEANUT","TREE_NUT",
        "SOY","WHEAT","GLUTEN","SESAME","LUPIN","SULPHITE"]
def srt(xs): return [a for a in ORDER if a in set(xs)]

matched, unmatched = 0, []
final = []
for p in proposals:
    name = p["name"]
    if name in R:
        c, m, conf, note = R[name]
        c = srt(c); m = srt([x for x in m if x not in c])
        final.append({**p, "contains": c, "mayContain": m, "confidence": conf,
                      "note": note, "source": "researched"})
        matched += 1
    else:
        if p["confidence"] == "search":
            unmatched.append(name)
        final.append({**p, "contains": srt(p["proposed"]), "mayContain": [],
                      "confidence": "high" if p["confidence"] == "high" else "low",
                      "note": p["reason"] or ("plain produce/herb/spirit - no allergens"
                                              if not p["proposed"] else ""),
                      "source": "rule"})

json.dump(final, open("/tmp/allergen-final-proposals.json", "w"), indent=2)

# ---- build xlsx ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Allergen proposals"
hdr = ["Ingredient", "Category", "Uses", "CONTAINS (propose to set)",
       "May contain (trace)", "Confidence", "Source", "Notes / verify"]
ws.append(hdr)
brown = PatternFill("solid", fgColor="7A3B2E")
amber = PatternFill("solid", fgColor="FFF2CC")
red = PatternFill("solid", fgColor="F8CBAD")
green = PatternFill("solid", fgColor="E2EFDA")
thin = Border(*[Side(style="thin", color="DDDDDD")] * 4)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = brown
    c.alignment = Alignment(vertical="center", wrap_text=True)

# sort: researched first, then rule-with-allergens, then rule-none
def keyf(x):
    rank = 0 if x["source"] == "researched" else (1 if x["contains"] else 2)
    return (rank, x["category"], x["name"].lower())
for x in sorted(final, key=keyf):
    ws.append([x["name"], x["category"], x["uses"], ", ".join(x["contains"]),
               ", ".join(x["mayContain"]), x["confidence"], x["source"], x["note"]])
    r = ws.max_row
    conf_cell = ws.cell(r, 6)
    if x["confidence"] == "low": conf_cell.fill = red
    elif x["confidence"] == "medium": conf_cell.fill = amber
    elif x["contains"]: conf_cell.fill = green
    for col in range(1, 9):
        ws.cell(r, col).border = thin
        ws.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)

widths = [34, 12, 6, 26, 28, 11, 11, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

out = "/Users/chris/C/tarte-kitchen/Allergen-Ingredient-Review-2026-06-24.xlsx"
wb.save(out)

# summary
researched = [x for x in final if x["source"] == "researched"]
ruled = [x for x in final if x["source"] == "rule"]
withc = [x for x in final if x["contains"]]
print(f"merged. matched-to-research: {matched}, unmatched search items: {unmatched}")
print(f"TOTAL empty ingredients: {len(final)}")
print(f"  -> propose CONTAINS allergens: {len(withc)}")
print(f"  -> propose NONE: {len(final)-len(withc)}")
print(f"researched: {len(researched)}  (low conf: {sum(1 for x in researched if x['confidence']=='low')}, medium: {sum(1 for x in researched if x['confidence']=='medium')}, high: {sum(1 for x in researched if x['confidence']=='high')})")
needs = [x for x in final if x["confidence"] in ("low","medium")]
print(f"items flagged for human/label check (low+medium conf): {len(needs)}")
print(f"\nwrote {out}")
